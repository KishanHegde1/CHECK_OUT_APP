"""In-memory CSV/XLSX student import with row-level savepoints."""

from __future__ import annotations

import csv
from collections.abc import Iterable
from io import BytesIO, StringIO
from pathlib import Path
from zipfile import BadZipFile

from fastapi import UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.core.exceptions import APIError
from app.schemas.student import (
    StudentImportError,
    StudentImportResult,
    StudentImportRow,
)
from app.services.admin_service import create_student

MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 1_000
REQUIRED_COLUMNS = frozenset(
    {
        "full_name",
        "email",
        "student_id",
        "department",
        "year",
        "room_number",
        "phone",
        "parent_phone",
        "hostel_status",
    }
)
FORBIDDEN_COLUMNS = frozenset(
    {
        "id",
        "user_id",
        "database_id",
        "password",
        "password_hash",
        "jwt",
        "token",
        "access_token",
    }
)
ALLOWED_CONTENT_TYPES = {
    ".csv": frozenset({"text/csv", "application/csv"}),
    ".xlsx": frozenset(
        {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
    ),
}


def _cell_value(value: object) -> str | None:
    """Convert spreadsheet values to trimmed strings without retaining files."""

    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _validate_headers(headers: Iterable[object]) -> list[str]:
    """Normalize and validate import headers before examining row contents."""

    normalized = [(_cell_value(header) or "").lower() for header in headers]
    if not normalized or any(not header for header in normalized):
        raise APIError(
            status_code=422, message="Import file has invalid column headers"
        )
    if len(set(normalized)) != len(normalized):
        raise APIError(
            status_code=422, message="Import file has duplicate column headers"
        )
    forbidden = sorted(FORBIDDEN_COLUMNS.intersection(normalized))
    if forbidden:
        raise APIError(
            status_code=422,
            message=f"Import file contains forbidden columns: {', '.join(forbidden)}",
        )
    missing = sorted(REQUIRED_COLUMNS.difference(normalized))
    if missing:
        raise APIError(
            status_code=422,
            message=f"Import file is missing required columns: {', '.join(missing)}",
        )
    return normalized


def _csv_rows(content: bytes) -> list[tuple[int, dict[str, str | None]]]:
    """Parse a UTF-8 CSV upload in memory."""

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise APIError(
            status_code=422, message="CSV file must use UTF-8 encoding"
        ) from exc
    reader = csv.DictReader(StringIO(text))
    headers = _validate_headers(reader.fieldnames or ())
    rows: list[tuple[int, dict[str, str | None]]] = []
    for row_number, row in enumerate(reader, start=2):
        normalized = {header: _cell_value(row.get(header)) for header in headers}
        if any(value is not None for value in normalized.values()):
            rows.append((row_number, normalized))
    return rows


def _xlsx_rows(content: bytes) -> list[tuple[int, dict[str, str | None]]]:
    """Parse the first worksheet of an XLSX upload in memory."""

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, OSError, ValueError) as exc:
        raise APIError(status_code=422, message="Invalid XLSX file") from exc
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = _validate_headers(next(rows, ()))
        parsed: list[tuple[int, dict[str, str | None]]] = []
        for row_number, values in enumerate(rows, start=2):
            normalized = {
                header: _cell_value(values[index] if index < len(values) else None)
                for index, header in enumerate(headers)
            }
            if any(value is not None for value in normalized.values()):
                parsed.append((row_number, normalized))
        return parsed
    finally:
        workbook.close()


async def import_students(
    db: Session,
    upload: UploadFile,
) -> StudentImportResult:
    """Import valid rows atomically per row and retain no uploaded file."""

    filename = upload.filename or ""
    extension = Path(filename).suffix.lower()
    allowed_types = ALLOWED_CONTENT_TYPES.get(extension)
    if allowed_types is None or upload.content_type not in allowed_types:
        await upload.close()
        raise APIError(
            status_code=422, message="Only CSV and XLSX student imports are allowed"
        )

    content = await upload.read(MAX_IMPORT_BYTES + 1)
    await upload.close()
    if len(content) > MAX_IMPORT_BYTES:
        raise APIError(status_code=422, message="Import file exceeds the 5 MB limit")

    rows = _csv_rows(content) if extension == ".csv" else _xlsx_rows(content)
    if len(rows) > MAX_IMPORT_ROWS:
        raise APIError(
            status_code=422, message="Import file exceeds the 1000-row limit"
        )

    errors: list[StudentImportError] = []
    seen_usernames: set[str] = set()
    seen_emails: set[str] = set()
    seen_student_ids: set[str] = set()
    imported = 0
    for row_number, raw_row in rows:
        try:
            row = StudentImportRow.model_validate(raw_row)
            username = row.full_name.casefold()
            email = row.email.casefold()
            student_id = row.student_id.casefold()
            if username in seen_usernames:
                raise APIError(
                    status_code=409, message="Duplicate full name in import file"
                )
            if email in seen_emails:
                raise APIError(
                    status_code=409, message="Duplicate email in import file"
                )
            if student_id in seen_student_ids:
                raise APIError(
                    status_code=409, message="Duplicate student ID in import file"
                )
            with db.begin_nested():
                create_student(db, row, commit=False)
            seen_usernames.add(username)
            seen_emails.add(email)
            seen_student_ids.add(student_id)
            imported += 1
        except ValidationError as exc:
            errors.append(
                StudentImportError(row=row_number, message=exc.errors()[0]["msg"])
            )
        except APIError as exc:
            errors.append(StudentImportError(row=row_number, message=exc.message))
        except SQLAlchemyError:
            errors.append(
                StudentImportError(
                    row=row_number, message="Database service is unavailable"
                )
            )

    try:
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise APIError(
            status_code=503, message="Database service is unavailable"
        ) from exc
    return StudentImportResult(
        total_rows=len(rows),
        imported=imported,
        failed=len(errors),
        errors=errors,
    )


__all__ = ["import_students"]
